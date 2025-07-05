from openpyxl import load_workbook
from openpyxl.styles import Font
import os

def generar_horario_docente(nombre_docente, turno, asignaciones):
    archivo_base = "FORMATO_HORARIO.xlsx"
    if not os.path.exists(archivo_base):
        print("❌ No se encontró el archivo de plantilla.")
        return

    wb = load_workbook(archivo_base)
    ws = wb.active  # Asumimos que la hoja activa es la correcta

    # Buscar la fila donde va el nombre del docente
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and "NOMBRE DEL PROFESOR" in str(cell.value).upper():
                cell.value = f"NOMBRE DEL PROFESOR: {nombre_docente}"
                cell.font = Font(bold=True)

    # Insertar asignaciones de prueba
    # Formato: (día, tramo, texto)
    for dia, tramo, texto in asignaciones:
        # Buscar la fila del tramo horario
        fila_tramo = None
        for row in ws.iter_rows(min_row=1, max_row=50):
            if row[0].value == tramo:
                fila_tramo = row
                break
        if not fila_tramo:
            print(f"⚠️ Tramo {tramo} no encontrado.")
            continue

        # Determinar columna según día
        dias_columnas = {
            "LUNES": 1,
            "MARTES": 2,
            "MIÉRCOLES": 3,
            "MIERCOLES": 3,
            "JUEVES": 4,
            "VIERNES": 5
        }

        col_idx = dias_columnas.get(dia.upper())
        if col_idx is not None:
            fila_tramo[col_idx].value = texto
            fila_tramo[col_idx].font = Font(name="Segoe UI", size=10)

    # Guardar como nuevo archivo
    nombre_archivo = f"Horario_{nombre_docente.replace(' ', '_')}.xlsx"
    wb.save(nombre_archivo)
    print(f"✅ Horario generado: {nombre_archivo}")